#  код (обёрнут в функцию run_analysis)
# os.environ['TCL_LIBRARY'] = r'C:\Program Files\Python313\tcl\tcl8.6'

import os
import traceback
import pandas as pd
import matplotlib
import json
import gc
import re

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from datetime import datetime
from pathlib import Path
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Border, Side

os.environ['TCL_LIBRARY'] = r'C:\Program Files\Python313\tcl\tcl8.6'


def extract_period_from_filename(file_path: Path) -> str:
    """
    Извлекает наименование периода из имени файла.
    Пример: 'Исходная таблица ЯНВАРЬ-АВГУСТ 25.xlsx' → 'ЯНВАРЬ-АВГУСТ 25'
    """
    match = re.search(r"Исходная таблица (.+?)\.xlsx", file_path.name, re.IGNORECASE)
    return match.group(1) if match else "Период не указан"


def log_message(msg: str):
    log_file = Path(__file__).resolve().parent / "error.log"
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(f"[{datetime.now().isoformat()}] {msg}\n")


def run_analysis(out_file: Path, input_file: Path, task_id: str, original_filename: Path):
# def run_analysis(out_file: Path, input_file: Path, task_id: str):
    try:
        start = datetime.now()
        log_message("=== Запуск анализа ===")

        # === 1. Загружаем конфиг ===
        BASE_DIR = Path(__file__).resolve().parent.parent
        CONFIG_PATH = BASE_DIR / "config.json"

        with open(CONFIG_PATH, encoding="utf-8") as f:
            config = json.load(f)

        abc_file = BASE_DIR / config["abc_file"]
        out_dir = out_file.parent
        xyz_thresholds = config["xyz_thresholds"]

        #period_days = config["period_days"]
        log_message(f"Определено X Y Z: {xyz_thresholds}")
        log_message("Конфиг загружен")

        # период оборачиваемости должен равняться количеству столбцов исходной таблицы без столбца Наименование
        # log_message(f"Определено X Y Z: {xyz_thresholds}")

        # пути промежуточных таблиц

        out_path_w = out_dir / "Исходная таблица_w.xlsx"
        out_path_ship = out_dir / "Исходная таблица_отгрузка.xlsx"
        out_path_stock = out_dir / "Исходная таблица_остаток.xlsx"

        # === 2. Трансформация исходной таблицы ===
        # df = pd.read_excel(input_file, header=None)
        log_message("Исходная таблица загружена")

        # -----------copilot----------------------------------------
        try:
            df = pd.read_excel(input_file, header=None)
        except Exception as e:
            log_message(f"❌ Ошибка чтения Excel: {e}")
            return
        # --------------copilot---------------------------------------

        # период оборачиваемости должен равняться количеству столбцов исходной таблицы без столбца Наименование
        period_days = df.shape[1] - 1  # Количество столбцов минус 1
        log_message(f"Определено количество дней: {period_days}")

        df.iloc[3, 1:] = df.iloc[3, 1:].replace("отгрузка продукта", "отгрузка")
        for i in range(4, len(df), 3):
            if i < len(df):
                df.iloc[i + 2, 0] = df.iloc[i, 0]
        df = df.drop(df.index[2::3]).reset_index(drop=True)
        df.to_excel(out_path_w, header=False, index=False)
        log_message("Таблица трансформирована и сохранена (W)")

        # === 3. Создание таблиц "отгрузка" и "остаток" ===
        df_ship = df.drop(df.index[1::2]).reset_index(drop=True)
        df_stock = df.drop(df.index[2::2]).reset_index(drop=True)

        def add_total_column(df_in, label: str):
            df_out = df_in.copy()
            df_num = df_out.iloc[:, 1:].replace("-", 0)
            df_num = pd.to_numeric(df_num.stack(), errors="coerce").unstack(fill_value=0)
            df_out["Всего"] = df_num.sum(axis=1).astype("object")
            if df_out.shape[0] > 0:
                df_out.at[0, "Всего"] = "Всего"
            if df_out.shape[0] > 1:
                df_out.at[1, "Всего"] = label
            return df_out

        df_ship = add_total_column(df_ship, "отгрузка")
        df_stock = add_total_column(df_stock, "остаток")

        df_stock_num = df_stock.iloc[:, 1:-1].replace("-", 0)
        df_stock_num = pd.to_numeric(df_stock_num.stack(), errors="coerce").unstack(fill_value=0)
        avg_values = df_stock_num.mean(axis=1).tolist()
        df_stock.insert(df_stock.shape[1], "Средний", avg_values)

        df_stock["Средний"] = df_stock["Средний"].astype("object")

        if df_stock.shape[0] > 0:
            df_stock.at[0, "Средний"] = "Средний"
        if df_stock.shape[0] > 1:
            df_stock.at[1, "Средний"] = "остаток"

        df_ship.to_excel(out_path_ship, header=False, index=False)
        df_stock.to_excel(out_path_stock, header=False, index=False)
        log_message("Файлы 'отгрузка' и 'остаток' сохранены")

        # ------------copilot------------------------------
        try:
            ship = pd.read_excel(out_path_ship, header=0)
            stock = pd.read_excel(out_path_stock, header=0)
            abc = pd.read_excel(abc_file)
        except Exception as e:
            log_message(f"❌ Ошибка чтения промежуточных файлов: {e}")
            return
        # -------------copilot----------------------------

        ship = ship[pd.to_numeric(ship["Всего"], errors="coerce").notna()]
        stock = stock[pd.to_numeric(stock["Средний"], errors="coerce").notna()]
        ship["Всего"] = pd.to_numeric(ship["Всего"], errors="coerce")
        stock["Средний"] = pd.to_numeric(stock["Средний"], errors="coerce")

        # ------------------------------------copilot-----------------
        if ship.empty or stock.empty:
            log_message("❌ ship или stock пусты после фильтрации")
            return
        #-----------------------------copilot------------------------

        df = pd.merge(
            ship[["Наименование", "Всего"]],
            stock[["Наименование", "Средний"]],
            on="Наименование",
            how="inner"
        )

        # ---------------------------------copilot---------------------
        if "Наименование" not in df.columns:
            log_message("❌ Нет столбца 'Наименование' после объединения")
            return
        # ----------copilot------------------------------------

        # добавляем ABC-группу
        df = pd.merge(df, abc[["Наименование", "Группа ABC"]], on="Наименование", how="left")

        # оборачиваемость
        df["Оборачиваемость_дни"] = df.apply(
            lambda x: (x["Средний"] * period_days) / x["Всего"]
            if pd.notna(x["Всего"]) and x["Всего"] > 0 else 9999,
            axis=1
        )

        # XYZ-группа
        def assign_xyz(turnover: float) -> str:
            if turnover <= xyz_thresholds["X"]:
                return "X"
            elif turnover <= xyz_thresholds["Y"]:
                return "Y"
            elif turnover <= xyz_thresholds["Z"]:
                return "Z"
            else:
                return "Неликвид"

        df["Оборачиваемость_дни"] = df["Оборачиваемость_дни"].round().astype(int)
        df["Группа XYZ"] = df["Оборачиваемость_дни"].apply(assign_xyz)
        df["ABC_XYZ"] = df["Группа ABC"] + "-" + df["Группа XYZ"]

        df.rename(columns={
            "Всего": "Всего отгрузка,кг",
            "Средний": "Средний остаток,кг"}, inplace=True)

        #============================================================
        # === Расширение данных: добавляем "Без группы" ===
        df_no_abc = df[df["Группа ABC"].isna() & df["Группа XYZ"].notna()].copy()
        df_no_abc["Группа ABC"] = "Без группы"
        df_no_abc["ABC_XYZ"] = df_no_abc["Группа ABC"] + "-" + df_no_abc["Группа XYZ"]

        df_full = pd.concat([df, df_no_abc], ignore_index=True)

        # === Сводная матрица по весу отгрузки ===
        pivot_weight = df_full.pivot_table(
            index="Группа ABC",
            columns="Группа XYZ",
            values="Всего отгрузка,кг",
            aggfunc="sum",
            fill_value=0
        )

        # Переводим в проценты
        total_weight = df_full["Всего отгрузка,кг"].sum()
        pivot_percent = (pivot_weight / total_weight * 100).round(2)

        # Переименовываем столбцы
        pivot_percent.columns = [f"{col}(%)" for col in pivot_percent.columns]

        # Проверка и создание файла, если он отсутствует
        if not out_file.exists():
            with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
                # Создаём пустой лист или записываем базовые данные
                pd.DataFrame().to_excel(writer, sheet_name="Инициализация", index=False)
            log_message(f"Создан новый Excel-файл: {out_file.name}")

        # === Сохраняем в Excel ===
        with pd.ExcelWriter(out_file, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df_full.to_excel(writer, sheet_name="ABC_XYZ_данные", index=False)
            pivot_percent.to_excel(writer, sheet_name="Сводная матрица")

        log_message("Сводная матрица с 'Без группы' и процентами по отгрузке сохранена")
        log_message("ABC-XYZ анализ рассчитан")

        # === Форматирование листа "ABC_XYZ_данные" ===
        wb = load_workbook(out_file)
        ws_data = wb["ABC_XYZ_данные"]

        # Заголовки — жирный шрифт и выравнивание
        for cell in ws_data[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Формат числовых колонок
        header_row = [cell.value for cell in ws_data[1]]
        numeric_columns = ["Всего отгрузка,кг", "Средний остаток,кг", "Оборачиваемость_дни"]

        for col_name in numeric_columns:
            if col_name in header_row:
                col_index = header_row.index(col_name) + 1
                for row in ws_data.iter_rows(min_row=2, min_col=col_index, max_col=col_index):
                    for cell in row:
                        cell.number_format = '0.00'

        # Автоширина столбцов
        for col in ws_data.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws_data.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        log_message("Форматирование листа 'ABC_XYZ_данные' завершено")
        wb.save(out_file)

        # ✅ Форматируем числовые значения до двух знаков после запятой
        df["Всего отгрузка,кг"] = df["Всего отгрузка,кг"].astype(float).round(2)
        df["Средний остаток,кг"] = df["Средний остаток,кг"].astype(float).round(2)

        # === Форматирование листа "Сводная матрица" ===
        # Переводим pivot_weight в доли, не проценты
        pivot_percent = (pivot_weight / total_weight).round(4)

        # Переименовываем столбцы
        pivot_percent.columns = [f"{col}(%)" for col in pivot_percent.columns]

        # Сохраняем в Excel
        with pd.ExcelWriter(out_file, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            pivot_percent.to_excel(writer, sheet_name="Сводная матрица")

        # Форматирование
        wb = load_workbook(out_file)
        ws_matrix = wb["Сводная матрица"]
        start_row = ws_matrix.max_row + 2

        ws_matrix.cell(row=start_row,
                column=1).value = " A – позиции УК, обеспечивающие 80% суммы М1 по факту продаж за 1 полугодие 2025г."
        ws_matrix.cell(row=start_row + 1,
                column=1).value = " B - позиции УК, обеспечивающие 15% суммы М1 по факту продаж за 1 полугодие 2025г."
        ws_matrix.cell(row=start_row + 2,
                column=1).value = " C - позиции УК, обеспечивающие 5% суммы М1 по факту продаж за 1 полугодие 2025г."
        ws_matrix.cell(row=start_row + 4,
                column=1).value = (" Оборачиваемость = Средний остаток товара на складе * Количество дней в периоде / "
                                   "Объем продаж (отгрузки) за  период")
        ws_matrix.cell(row=start_row + 6,
                       column=1).value = " Х <= 30; Y <= 60; Z <= 90; Неликвид > 90"

        # Заголовки
        for cell in ws_matrix[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Формат процентов
        for row in ws_matrix.iter_rows(min_row=2, min_col=2):
            for cell in row:
                cell.number_format = '0.00%'

        # Цвет строки "Без группы"
        fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        for row in ws_matrix.iter_rows(min_row=2, max_row=ws_matrix.max_row):
            if row[0].value == "Без группы":
                for cell in row:
                    cell.fill = fill

        # Автоширина
        for col in ws_matrix.columns:
            col_letter = get_column_letter(col[0].column)
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)

            # Если это первый столбец (Группа ABC) — уменьшаем ширину в 3 раза
            if col[0].column == 1:
                ws_matrix.column_dimensions[col_letter].width = max(max_length // 3, 5)
            else:
                ws_matrix.column_dimensions[col_letter].width = max_length + 2

        # добавить тонкие границы только к строкам
        # "Группа ABC", "A", "B", "C", "Без группы" в листе "Сводная матрица"
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        target_rows = {"Группа ABC", "A", "B", "C", "Без группы"}

        for row in ws_matrix.iter_rows(min_row=2, max_row=ws_matrix.max_row):
            row_label = str(row[0].value).strip() if row[0].value else ""
            if row_label in target_rows:
                for cell in row:
                    cell.border = thin_border

        wb.save(out_file)
        log_message("Форматирование 'Сводной матрицы' завершено")

        # === Удаление лишнего листа "Инициализация" ===
        if "Инициализация" in wb.sheetnames:
            wb.remove(wb["Инициализация"])
            log_message("Удалён лишний лист 'Инициализация'")

        wb.save(out_file)

        # Группировка по ABC_XYZ
        weights = df_full.groupby("ABC_XYZ")["Всего отгрузка,кг"].sum()
        weights_percent = (weights / total_weight * 100).round(2)

        xyz_info = (
            f"X ≤ {xyz_thresholds['X']} дн., "
            f"Y ≤ {xyz_thresholds['Y']} дн., "
            f"Z ≤ {xyz_thresholds['Z']} дн."
        )

        # определение периода ABC-XYZ анализа  из original_filename
        period_name = extract_period_from_filename(original_filename)
        if period_name == "Период не указан":
            log_message("⚠️ Не удалось извлечь период из имени файла. Используется fallback.")
        log_message(f"📅 Период анализа: {period_name}")

        # горизонтальная столбчатая диаграмма
        weights_percent.sort_values().plot.barh(
            figsize=(10, 8),
            color="skyblue",
            edgecolor="black"
        )
        plt.xlabel("Доля отгрузки, %")
        # plt.title(f"ABC-XYZ анализ январь_август 25\n{xyz_info}\nПериод: {period_days}", fontsize=11)

        plt.title(f"ABC-XYZ анализ {period_name}\n{xyz_info}\nПериод: {period_days}", fontsize=11)
        plt.tight_layout()

        chart_path = out_dir / f"ABC_XYZ_график_{period_name}.png"
        plt.savefig(chart_path, dpi=300)
        plt.close()
        log_message("Диаграмма сохранена как PNG")

        #  ========================================
        # встраиваем диаграмму в Excel
        wb = load_workbook(out_file)
        ws_chart = wb["Диаграмма"] if "Диаграмма" in wb.sheetnames else wb.create_sheet(f"Диаграмма {period_name}")

        img = Image(str(chart_path))
        img.width, img.height = 480, 480
        ws_chart.add_image(img, "B2")
        # =========================================

        wb.save(out_file)

        log_message("Диаграмма добавлена в Excel")
        log_message("=== Анализ успешно завершён ===")
        # -------------copilot-------------------
        log_message(f"✅ Анализ завершён: {out_file.name}")
        log_message(f"⏱️ Время выполнения: {(datetime.now() - start).total_seconds():.2f} сек")
        gc.collect()
        # --------copilot------------------------------

    except Exception as e:
        error_log = Path(__file__).resolve().parent / "error.log"
        with open(error_log, "a", encoding="utf-8") as f:
            f.write(f"\n[ERROR] Ошибка при выполнении анализа: {e}\n")
            f.write(traceback.format_exc() + "\n")





